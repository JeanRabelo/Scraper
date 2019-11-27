import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from pprint import pprint

PATH_IN = r'C:\Users\jean_\Documents\GitHub\Up.p\Operacional\cadastro_positivo\consolidado\021\dados\021.xlsx'

wb = load_workbook(filename = PATH_IN)

parent_map = dict()

def printar_elem_corrente(wb, i, row, col):
    print('Nome completo = ' + str(wb.worksheets[i].cell(row, col).value), end=' ')
    print('Nome reduzido = ' + str(wb.worksheets[i].cell(row, col + 1).value), end=' ')
    print('Valor = ' + str(wb.worksheets[i].cell(row, col + 2).value))

def fill_elem(elem, wb, i, row, col):
    elem.set(str(wb.worksheets[i].cell(row, col + 1).value), str(wb.worksheets[i].cell(row, col + 2).value))

root = ET.Element('EnvoCad')
current_element = root

index_curr_sheet = 0

element_level = 0
n_sheets = len(wb.sheetnames)
row = 2
col = 1

for i in range(0,n_sheets):
    while wb.worksheets[i].cell(row, col).value != None or wb.worksheets[i].cell(1, col + 4).value != None:
        while wb.worksheets[i].cell(row, col).value != None:
            if str(wb.worksheets[i].cell(row, col + 2).value)[:10] == r'tag_level=':
                current_level = int(str(wb.worksheets[i].cell(row, col + 2).value)[10:])
                if current_level <= element_level:
                    for j in range(0, element_level - current_level + 1):
                        current_element = parent_map[current_element]

                next_current_element = ET.SubElement(current_element, str(wb.worksheets[i].cell(row, col + 1).value))
                parent_map[next_current_element] = current_element
                current_element = next_current_element
                element_level = current_level
            else:
                fill_elem(current_element, wb, i, row, col)

            printar_elem_corrente(wb, i, row, col)
            row = row + 1

        if wb.worksheets[i].cell(1, col + 4).value != None:
            print('Trocou de coluna')
            row = 1
            col = col + 4

    print('Trocou de worksheet')
    row = 1
    col = 1

print('Acabou!')
tree = ET.ElementTree(root)
PATH_OUT = r'C:\Users\jean_\Documents\GitHub\Up.p\Operacional\cadastro_positivo\consolidado\021\resultado\BVXA021_34046974_11725176_20191118_00001.xml'
tree.write(PATH_OUT)
