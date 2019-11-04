import xml.etree.ElementTree as ET
from openpyxl import load_workbook

PATH_IN = r'consolidado\005\dados\005.xlsx'
PATH_OUT = r'consolidado\005\resultado\BVXA005_34046974_11725176_20191002_00001.xml'
SET_DIRECAO = {'Abaixo', 'Acima', 'Lado'}


root = ET.Element('EnvoHstCrd')
element_atual = root

# PATH_IN = r'C:\Users\jean_\Documents\GitHub\Up.p\Operacional\cadastro_positivo\consolidado\005\dados\005.xlsx'
# PATH_OUT = r'C:\Users\jean_\Documents\GitHub\Up.p\Operacional\cadastro_positivo\consolidado\005\resultado\BVXA005_34046974_11725176_20191002_00001.xml'

wb = load_workbook(filename = PATH_IN)
index_sheet_atual = 0
row_atual = 2
col_atual = 1
element_level = 1

def parent_node(root, n_volta, elemento):
    tree = ET.ElementTree(root)
    parent_map = {c:p for p in tree.iter() for c in p}
    for i in rante(0, n_volta):
        elemento = parent_map[parent_map]

    return = elemento


def atualizar_posicao(cell_atual):
    return [cell_atual.value is Null,
    str(cell_atual.offset(0, 2).value)[:10] == r'tag_level=']

[eh_fim_de_coluna, eh_tag] = atualizar_posicao(wb[index_sheet_atual].cell(row_atual, col_atual))

while not eh_fim_de_sheet:
    while not eh_fim_de_coluna:
        while not eh_fim_de_coluna and not eh_tag: # Preencher atributos da tag atual
            atributo_atual = wb[index_sheet_atual].cell(row_atual, col_atual + 1).value
            valor_atributo_atual = wb[index_sheet_atual].cell(row_atual, col_atual + 2).value
            element_atual.set(atributo_atual, valor_atributo_atual)
            row_atual = row_atual + 1
            [eh_fim_de_coluna, eh_tag] = atualizar_posicao(wb[index_sheet_atual].cell(row_atual, col_atual))

        if eh_tag: # Criar elemento adequado
            nivel_alcancado = int(str(wb[index_sheet_atual].cell(row_atual, col_atual + 2).value)[10:])
            # Se nível a que chegamos é maior (mais profundo) que o anterior:
            if nivel_alcancado > element_level:
                element_atual = ET.SubElement(element_atual, wb[index_sheet_atual].cell(row_atual, col_atual + 1))
            elif nivel_alcancado < element_level:
                element_atual = parent_node(root=root, n_volta=(element_level-nivel_alcancado), element_atual)
            else:
                element_atual = parent_node(root=root, n_volta=1, element_atual)
                element_atual = ET.SubElement(element_atual, wb[index_sheet_atual].cell(row_atual, col_atual + 1))

            row_atual = row_atual + 1
            element_level = nivel_alcancado

    [eh_fim_de_coluna, eh_tag] = atualizar_posicao(wb[index_sheet_atual].cell(row_atual, col_atual))



tree = ET.ElementTree(root)
tree.write(PATH_OUT)
