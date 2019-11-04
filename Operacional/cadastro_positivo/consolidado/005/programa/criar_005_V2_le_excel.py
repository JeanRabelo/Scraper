import xml.etree.ElementTree as ET
from openpyxl import load_workbook

PATH_IN = r'C:\Users\jean_\Documents\GitHub\Up.p\Operacional\cadastro_positivo\consolidado\005\dados\005.xlsx'
PATH_OUT = r'C:\Users\jean_\Documents\GitHub\Up.p\Operacional\cadastro_positivo\consolidado\005\resultado\BVXA005_34046974_11725176_20191002_00001.xml'

wb = load_workbook(filename = PATH_IN)

root = ET.Element('EnvoHstCrd')
current_element = root

index_curr_sheet = 0

element_level = 0
n_sheets = len(wb.sheetnames)
row = 1
col = 1

for i in range(0,n_sheets):
    while wb.worksheets[i].cell(row, col).value != None or wb.worksheets[i].cell(1, col + 4).value != None:
        while wb.worksheets[i].cell(row, col).value != None:
            print('Nome completo = ' + str(wb.worksheets[i].cell(row, col).value), end=' ')
            print('Nome reduzido = ' + str(wb.worksheets[i].cell(row, col + 1).value), end=' ')
            print('Valor = ' + str(wb.worksheets[i].cell(row, col + 2).value))
            input('aaaa')
            row = row + 1

        if wb.worksheets[i].cell(1, col + 4).value != None:
            print('Trocou de coluna')
            input('aaaa')
            row = 1
            col = col + 4

    print('Trocou de worksheet')
    row = 1
    col = 1


# tree = ET.ElementTree(root)
# tree.write(PATH_OUT)
